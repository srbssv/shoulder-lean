from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import os
import csv

# Этот скрипт формирует отчет о текущем составе классов

fieldnames = ['GivenName', 'Surname', 'SamAccountName', 'UserPrincipalName']
xlfieldnames = ['Имя', 'Фамилия', 'ИИН', 'Почта']
ft_header = Font(size=12, bold=True)
ft_table_header = Font(size=12, bold=True)
ft1 = Font(size=11)
ali = Alignment(horizontal="center", vertical="center")
sid = Side(border_style="thin", color="000000")
bord = Border(sid, sid, sid, sid)
fill = PatternFill("solid", fgColor="cfcfcf")

# Берем все csv-файлы в окрестности
files = [f for f in os.listdir('.') if os.path.splitext(f)[1] == '.csv']
for f in files:
    classname = f.split(".")[0]
    xl = Workbook()
    sheet = xl.active
    sheet.cell(1, 1, classname)
    sheet.append(xlfieldnames)

    # Читаем оттуда списки учеников
    with open(f, 'r', newline='', encoding='utf8') as csvfile:
        r = csv.reader(csvfile, delimiter=';')
        for row in r:
            if row[1] != 'Surname':
                sheet.append(row)
    sheet['A1'].font = ft_header
    sheet['A1'].alignment = ali

    # И записываем в excel-файл в папке classes
    for i in range(2, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if i == 2:
                sheet.cell(i, j).font = ft_table_header
                sheet.cell(i, j).alignment = ali
                sheet.cell(i, j).fill = fill
            else:
                sheet.cell(i, j).font = ft1
                sheet.cell(i, j).border = bord

# Ширину ячеек делаем по длине слов
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        col = chr(col + 64)      # Работает после openpyxl 2.6-a1 версии. При более старой версии эту строчку нужно закомментить. Также не работает для столбцов с названием 'AA', 'AB', и т.д.
        sheet.column_dimensions[col].width = value + 2      # Если версия openpyxl < 2.6-a1, закомментируйте предыдущую строчку
    sheet.merge_cells('A1:D1')
    fn = "classes/" + classname + ".xlsx"
    xl.save(fn)
    xl.close()

print("Классы успешно сохранены в папке classes/")
input("Нажми Enter")
